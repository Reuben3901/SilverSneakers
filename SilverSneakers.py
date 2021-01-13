#! python3
import openpyxl, pprint

print('Building Dictionaries...')

#------------------------------------------
#---------Create Processing Log------------
#------------------------------------------

processLog = open('Processing_Log.txt', 'w')
processLog.write('Creating Dictionaries\n')
processLog.close()
processLog = open('Processing_Log.txt', 'a')

#------------------------------------------
#----Create Front Door Keys Dictionary-----
#------------------------------------------

ml = openpyxl.load_workbook('CompleteMemberList.xlsx')
sheet = ml.active
keysDict = {}

for row in range(1, sheet.max_row + 1):
    SS    = sheet['A' + str(row)].value
    Last  = sheet['B' + str(row)].value
    First = sheet['C' + str(row)].value

    keysDict.setdefault(SS, {})
    keysDict[SS]['Last'] = str(Last)
    keysDict[SS]['First'] = str(First)

processLog.write('Created Front Door Dictionary\n')

#------------------------------------------
#--------Create MindBody Dictionary--------
#------------------------------------------

MindBodyDict = {}

for row in range(1, sheet.max_row + 1):
    SS  = sheet['A' + str(row)].value
    Last = sheet['B' + str(row)].value
    First    = sheet['C' + str(row)].value

    # Make sure the key for Last Name exists.
    MindBodyDict.setdefault(Last, {})
    MindBodyDict[Last].setdefault(First, {'SS': 0})
    MindBodyDict[Last][First]['SS'] = str(SS)

processLog.write('Created MindBody Dictionary\n')

#------------------------------------------
#--------Process Front Door Report---------
#------------------------------------------

processLog.write('Processing FrontDoorKeysReport.txt \n')

import re

wb = openpyxl.Workbook()
sheet = wb.active

ReportRow = 1

SSRegex = re.compile("(\d{5,})|\w(\d{5,})")
DateRegex = re.compile("(\d*\/\d*\/\d*)") #date
TimeRegex = re.compile("(\d*\:\d*\:\d*)") #time

with open("FrontDoorKeysReport.txt") as f:
    for line in f:
        SS = SSRegex.search(line)
        Date = DateRegex.search(line)
        Time = TimeRegex.search(line)
        if SS != None:
            try:
                sheet['A' + str(ReportRow)].value = keysDict[SS.group()]['Last']
                sheet['B' + str(ReportRow)].value = keysDict[SS.group()]['First']
                sheet['C' + str(ReportRow)].value = SS.group()
                sheet['D' + str(ReportRow)].value = Date.group()
                sheet['E' + str(ReportRow)].value = Time.group()
                ReportRow += 1
            except:
                processLog.write('\n--Warning! Bad Record Found! \n')
                processLog.write(str(SS.group()) + ' is not a valid SS number, compare records on MindBody with KERI DOORS\n\n')

#------------------------------------------
#---------Process MindBody Report----------
#------------------------------------------

print('Processing MindBodyReport.xlsx...')

mb = openpyxl.load_workbook('MindBodyReport.xlsx')
mbSheet = mb.active

processLog.write('Processing MindBodyReport.xlsx...\n')

DateRegex = re.compile("(\d+\/\d+\/\d+)") #date
TimeRegex = re.compile("(\d*(\:\d*)*) [aApPmM]*") #time
NameRegex = re.compile("[a-zA-Z-.]+") #name
CommaRegex = re.compile("[,]") #name

for SearchRow in range(1, mbSheet.max_row + 1):
    Date = DateRegex.search(str(mbSheet['A' + str(SearchRow)].value))
    Comma = CommaRegex.search(str(mbSheet['A' + str(SearchRow)].value))

    if Comma != None:
        Name = NameRegex.findall(str(mbSheet['A' + str(SearchRow)].value))

        try:
            gotdata = Name[2]
        except:
            gotdata = 'null'

        if gotdata != 'null':
            sheet['A' + str(ReportRow)].value = (str(Name[0]) + ' ' + str(Name[1]))
            sheet['B' + str(ReportRow)].value = str(Name[2])
            try:
                sheet['C' + str(ReportRow)].value = MindBodyDict[sheet['A' + str(ReportRow)].value][sheet['B' + str(ReportRow)].value]['SS']
            except:
                if str(Name[0]) != 'HYPERLINK':
                    processLog.write('\n--Warning! Bad Record Found! \n')
                    processLog.write('Last: ' + str(Name[0]) +' First: ' + str(Name[1]) + ' is invalid. Check FIRST & LAST & NUMBER on MindBody/Complete Silver Sneakers list!!\n\n')
        else:
            sheet['A' + str(ReportRow)].value = str(Name[0])
            sheet['B' + str(ReportRow)].value = str(Name[1])
            try:
                sheet['C' + str(ReportRow)].value = MindBodyDict[sheet['A' + str(ReportRow)].value][sheet['B' + str(ReportRow)].value]['SS']
            except:
                processLog.write('\n--Warning! Bad Record Found! \n')
                processLog.write('Last: ' + str(Name[0]) + ' First: ' + str(Name[1]) + ' is invalid. Check FIRST & LAST & NUMBER on MindBody/Complete Silver Sneakers list!!\n\n')
        ReportRow += 1
        
    if Date != None:
        sheet['D' + str(ReportRow)].value =str(Date.group(0))
        sheet['E' + str(ReportRow)].value =(str(mbSheet['B' + str(SearchRow)].value))

    SearchRow += 1

#------------------------------------------
#--------Remove Duplicates-----------------
#------------------------------------------

processLog.write('Cleaning up, removing duplicates!\n')

iterations = 0
duplicateCounter = 0

#Clean up the date 01/01/17 = 1/1/17
for row in range(1, sheet.max_row + 1):
    Date1 = str(sheet['D' + str(row)].value)
    if Date1[0] == '0' and Date1[3] == '0':
        sheet['D' + str(row)].value =str(Date1[1:3]+Date1[4:10])
    Date1 = str(sheet['D' + str(row)].value)
    if Date1[0] == '0':
        sheet['D' + str(row)].value =str(Date1[1:10])

print('Removing Duplicate Check-ins...')

for row in range(2, sheet.max_row + 1):
    x = (row/sheet.max_row * 100)
    if (10 < x < 10.05):
        print('10% Complete')
    if (20 < x < 20.05):
        print('20% Complete')
    if (30 < x < 30.05):
        print('30% Complete')
    if (40 < x < 40.05):
        print('40% Complete')
    if (50 < x < 50.05):
        print('50% Complete')
    if (60 < x < 60.05):
        print('60% Complete')
    if (70 < x < 70.05):
        print('70% Complete')
    if (80 < x < 80.05):
        print('80% Complete')
    if (90 < x < 90.05):
        print('90% Complete')

    SS1  = sheet['C' + str(row)].value
    Date1    = sheet['D' + str(row)].value
    for check in range((row+1), sheet.max_row + 1):
        SS2  = sheet['C' + str(check)].value
        Date2    = sheet['D' + str(check)].value
        if SS1 == SS2 and Date1 == Date2:
            sheet['A' + str(check)].value = str(' ')
            sheet['B' + str(check)].value = str(' ')
            sheet['C' + str(check)].value = str(' ')
            sheet['D' + str(check)].value = str(' ')
            sheet['E' + str(check)].value = str(' ')
            if SS1[0].isdecimal() == True:
                duplicateCounter += 1

wb = openpyxl.Workbook()
Sheet = wb.active

print('Removing Empty Rows...')

counter = 0
for row in range(1, sheet.max_row + 1):
    if str(sheet['D' + str(row)].value) != 'None' and str(sheet['D' + str(row)].value) != ' ':
        counter += 1
        Sheet['A' + str((counter))].value = sheet['A' + str(row)].value
        Sheet['B' + str((counter))].value = sheet['B' + str(row)].value
        Sheet['C' + str((counter))].value = sheet['C' + str(row)].value
        Sheet['D' + str((counter))].value = sheet['D' + str(row)].value
        Sheet['E' + str((counter))].value = sheet['E' + str(row)].value


processLog.write(str(duplicateCounter) + ' Duplicates deleted!\n')

#------------------------------------------
#-------------Split Reports----------------
#------------------------------------------


wb1 = openpyxl.Workbook()
wb2 = openpyxl.Workbook()
wb3 = openpyxl.Workbook()
Sheet1 = wb1.active
Sheet2 = wb2.active
Sheet3 = wb3.active

print('Splitting Reports...')

counter1 = 0
counter2 = 0
counter3 = 0

for row in range(1, Sheet.max_row + 1):
    if len(str(Sheet['C' + str(row)].value)) == 8:
        counter1 += 1
        Sheet1['A' + str((counter1))].value = Sheet['A' + str(row)].value
        Sheet1['B' + str((counter1))].value = Sheet['B' + str(row)].value
        Sheet1['C' + str((counter1))].value = Sheet['C' + str(row)].value
        Sheet1['D' + str((counter1))].value = Sheet['D' + str(row)].value
        Sheet1['E' + str((counter1))].value = Sheet['E' + str(row)].value
    if len(str(Sheet['C' + str(row)].value)) == 10:
        counter2 += 1
        Sheet2['A' + str((counter2))].value = Sheet['A' + str(row)].value
        Sheet2['B' + str((counter2))].value = Sheet['B' + str(row)].value
        Sheet2['C' + str((counter2))].value = Sheet['C' + str(row)].value
        Sheet2['D' + str((counter2))].value = Sheet['D' + str(row)].value
        Sheet2['E' + str((counter2))].value = Sheet['E' + str(row)].value
    if len(str(Sheet['C' + str(row)].value)) == 16:
        counter3 += 1
        Sheet3['A' + str((counter3))].value = Sheet['A' + str(row)].value
        Sheet3['B' + str((counter3))].value = Sheet['B' + str(row)].value
        Sheet3['C' + str((counter3))].value = Sheet['C' + str(row)].value
        Sheet3['D' + str((counter3))].value = Sheet['D' + str(row)].value
        Sheet3['E' + str((counter3))].value = Sheet['E' + str(row)].value
        
wb1.save('Silver&Fit.xlsx')
wb2.save('OptumFitness.xlsx')
wb3.save('SilverSneakers.xlsx')

#------------------------------------------
#---------Calculate Statistics-------------
#------------------------------------------

print('Calculating Statistics...')

ssc = openpyxl.load_workbook('SilverSneakers.xlsx')
sheet = ssc.active

uniqueVisitors = []
for row in range(1, sheet.max_row + 1):
    if str(sheet['C' + str(row)].value) not in uniqueVisitors:
        if str(sheet['C' + str(row)].value) != 'None' and str(sheet['C' + str(row)].value) != ' ':
            uniqueVisitors = uniqueVisitors + [sheet['C' + str(row)].value] # list concatenation

PaidCounter = 0
UniqueCounter = 0
TotalCounter = 0

for i in range(len(uniqueVisitors)):
    for row in range(1, sheet.max_row + 1):
        if uniqueVisitors[i] == sheet['C' + str(row)].value:
            UniqueCounter += 1
            TotalCounter += 1
    if UniqueCounter > 10:
        UniqueCounter = 10
    PaidCounter += UniqueCounter
    UniqueCounter = 0
            
processLog.write(str(len(uniqueVisitors)) + ' Unique Silver Sneaker Members\n')

processLog.write(str(TotalCounter) + ' Total Silver Sneaker Member Visits\n')
processLog.write(str(PaidCounter) + ' Paid Silver Sneaker Member Visis\n\n')
processLog.write('ALL DONE!\n')
processLog.write('Double Check Excel Records.........')
processLog.close()
